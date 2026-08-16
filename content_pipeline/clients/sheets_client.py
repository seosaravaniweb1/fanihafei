"""لایه‌ی خروجی: گوگل‌شیت و فایل محلی.

اصل ۵ معماری: گوگل‌شیت فقط لایه‌ی خروجی است، نه لایه‌ی کاری. برای فرار از
rate limit (حدود ۶۰ نوشتن در دقیقه) فقط یک batch write در انتها انجام می‌شود.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence


@dataclass
class Sheet:
    """یک تب خروجی."""

    title: str
    header: list[str]
    rows: list[list[Any]]


class SheetsWriter:
    """نوشتن در Google Sheets با ``gspread`` — فقط یک batch در پایان."""

    def __init__(self, sheet_id: str, service_account_json: str) -> None:
        self.sheet_id = sheet_id
        self.service_account_json = service_account_json

    @property
    def available(self) -> bool:
        if not (self.sheet_id and self.service_account_json):
            return False
        try:  # pragma: no cover - وابسته به محیط
            import importlib

            importlib.import_module("gspread")
        except ImportError:
            return False
        return Path(self.service_account_json).exists()

    def write(self, sheets: Sequence[Sheet]) -> str:  # pragma: no cover - نیازمند شبکه
        import gspread  # type: ignore

        client = gspread.service_account(filename=self.service_account_json)
        spreadsheet = client.open_by_key(self.sheet_id)
        for sheet in sheets:
            values = [sheet.header] + [[_cell(v) for v in row] for row in sheet.rows]
            try:
                worksheet = spreadsheet.worksheet(sheet.title)
                worksheet.clear()
            except gspread.WorksheetNotFound:
                worksheet = spreadsheet.add_worksheet(
                    title=sheet.title,
                    rows=max(len(values) + 10, 100),
                    cols=max(len(sheet.header) + 2, 12),
                )
            # یک کال به ازای هر تب — کل داده یکجا
            worksheet.update(values, "A1", value_input_option="RAW")
        return f"https://docs.google.com/spreadsheets/d/{self.sheet_id}"


def write_xlsx(path: str | Path, sheets: Sequence[Sheet]) -> str:
    """خروجی محلی. اگر ``openpyxl`` نبود، به چند فایل CSV برمی‌گردیم."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        return _write_csv_bundle(path, sheets)

    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet in sheets:
        worksheet = workbook.create_sheet(title=sheet.title[:31])
        worksheet.sheet_view.rightToLeft = True
        worksheet.append(sheet.header)
        for row in sheet.rows:
            worksheet.append([_cell(v) for v in row])
        for index, column in enumerate(sheet.header, start=1):
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=index).column_letter
            ].width = max(18, min(60, len(str(column)) + 12))
    workbook.save(path)
    return str(path)


def _write_csv_bundle(path: Path, sheets: Sequence[Sheet]) -> str:
    written: list[str] = []
    for sheet in sheets:
        target = path.with_name(f"{path.stem}__{_slug(sheet.title)}.csv")
        with target.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(sheet.header)
            for row in sheet.rows:
                writer.writerow([_cell(v) for v in row])
        written.append(str(target))
    return ", ".join(written)


def _cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return "، ".join(str(v) for v in value)
    if isinstance(value, bool):
        return "بله" if value else "خیر"
    return value


def _slug(text: str) -> str:
    return "".join(ch if ch.isalnum() else "-" for ch in text).strip("-") or "sheet"
