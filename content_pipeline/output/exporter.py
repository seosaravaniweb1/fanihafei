"""خروجی نهایی — سه شیت در یک فایل (اکسل) یا یک گوگل‌شیت.

* **شیت A — «آماده تولید محتوا»** (``has_suggest``)
* **شیت B — «آرشیو آینده»** (``no_suggest``) — همان ستون‌ها بدون ``lsi_keywords``
* **شیت C — «نیاز به بازبینی دستی»** — ادغام‌های مبهم + عنوان‌های مرزی موضوعی

گوگل‌شیت فقط لایه‌ی خروجی است: یک batch write در انتها (نه در طول کار).
"""

from __future__ import annotations

import csv
import io
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence

from ..core import db
from ..core.config import Config
from ..phases.p3_suggest import HAS_SUGGEST, NO_SUGGEST

READY_HEADER = [
    "canonical_title",
    "lsi_keywords",
    "lsi_count",
    "category",
    "source_domains",
    "source_urls",
    "merge_count",
    "confidence",
]

ARCHIVE_HEADER = [
    "canonical_title",
    "lsi_count",
    "category",
    "source_domains",
    "source_urls",
    "merge_count",
    "confidence",
]

# شیت «همه محصولات»: همه در یک جا، با ستونی که می‌گوید کدام ساجست دارد.
ALL_HEADER = [
    "canonical_title",
    "suggest",
    "lsi_keywords",
    "lsi_count",
    "category",
    "source_domains",
    "source_urls",
    "merge_count",
    "confidence",
]

SUGGEST_LABEL = {HAS_SUGGEST: "دارد", NO_SUGGEST: "ندارد", "": "بررسی نشده"}

# کلید هر شیت، برای انتخاب در config و پنل
SHEET_KEYS = ("ready", "archive", "all", "review")
DEFAULT_SHEETS = list(SHEET_KEYS)

REVIEW_HEADER = [
    "type",
    "title",
    "reason",
    "score",
    "source_urls",
]


@dataclass
class Sheet:
    """یک تب خروجی."""

    title: str
    header: list[str]
    rows: list[list[Any]]


@dataclass
class ExportStats:
    ready_rows: int = 0
    archive_rows: int = 0
    all_rows: int = 0
    review_rows: int = 0
    local_path: str = ""
    sheet_url: str = ""

    def render(self) -> str:
        parts = [
            f"خروجی — آماده تولید محتوا: {self.ready_rows}، "
            f"آرشیو آینده: {self.archive_rows}، همه محصولات: {self.all_rows}، "
            f"نیاز به بازبینی: {self.review_rows}"
        ]
        if self.local_path:
            parts.append(f"فایل محلی: {self.local_path}")
        if self.sheet_url:
            parts.append(f"گوگل‌شیت: {self.sheet_url}")
        return "\n".join(parts)


# ---------------------------------------------------------------------------
# ساخت ردیف‌ها
# ---------------------------------------------------------------------------


def selected_sheets(config: Config) -> list[str]:
    """کدام شیت‌ها ساخته شوند. ``output.sheets`` در config یا انتخاب پنل."""
    raw = config.get("output.sheets", DEFAULT_SHEETS) or DEFAULT_SHEETS
    if isinstance(raw, str):
        raw = [part.strip() for part in raw.split(",")]
    chosen = [key for key in SHEET_KEYS if key in {str(item).strip() for item in raw}]
    return chosen or DEFAULT_SHEETS


def build_one(conn: sqlite3.Connection, run_id: str, config: Config, key: str) -> Sheet:
    """فقط یک شیت — برای دانلود جداگانه‌ی «با ساجست» یا «بدون ساجست»."""
    if key not in SHEET_KEYS:
        raise ValueError(f"نام خروجی نامعتبر: {key}")
    saved = config.get("output.sheets")
    config.raw.setdefault("output", {})["sheets"] = [key]
    try:
        return build_sheets(conn, run_id, config)[0]
    finally:
        if saved is None:
            config.raw["output"].pop("sheets", None)
        else:
            config.raw["output"]["sheets"] = saved


def build_sheets(conn: sqlite3.Connection, run_id: str, config: Config) -> list[Sheet]:
    tabs = config.get("output.tabs", {}) or {}
    wanted = selected_sheets(config)
    ready: list[list] = []
    archive: list[list] = []
    everything: list[list] = []
    review: list[list] = []

    for row in db.canonical_products(conn, run_id):
        canonical_id = int(row["id"])
        keywords = [k["keyword"] for k in db.keywords_for(conn, canonical_id)]
        domains = db.source_domains_for(conn, canonical_id)
        urls = db.source_urls_for(conn, canonical_id)
        confidence = round(float(row["merge_confidence"] or 0.0), 3)
        category = db.category_of(conn, canonical_id)
        status = row["suggest_status"] or ""
        joined_keywords = " | ".join(keywords)
        tail = [category, " | ".join(domains), " | ".join(urls), len(urls), confidence]

        everything.append(
            [
                row["canonical_title"],
                SUGGEST_LABEL.get(status, "بررسی نشده"),
                joined_keywords,
                len(keywords),
                *tail,
            ]
        )
        if status == HAS_SUGGEST:
            ready.append([row["canonical_title"], joined_keywords, len(keywords), *tail])
        else:
            archive.append([row["canonical_title"], len(keywords), *tail])

        if row["needs_review"]:
            review.append(
                [
                    "ادغام مبهم",
                    row["canonical_title"],
                    "ابزار مطمئن نبود این عنوان باید با عنوان مشابهی ادغام شود یا نه",
                    confidence,
                    " | ".join(urls),
                ]
            )

    for raw in db.borderline_raw_products(conn, run_id):
        review.append(
            [
                "ارتباط مرزی با موضوع",
                raw["raw_title"],
                "امتیاز موضوعی بین آستانه‌ی رد و آستانه‌ی پذیرش",
                round(float(raw["topic_score"] or 0.0), 3),
                raw["source_url"],
            ]
        )

    # پرمحتواترین‌ها اول، تا اولویت تولید محتوا از بالای شیت مشخص باشد
    ready.sort(key=lambda r: (-r[2], r[0]))
    archive.sort(key=lambda r: r[0])
    everything.sort(key=lambda r: (r[1] != "دارد", -r[3], r[0]))
    review.sort(key=lambda r: (r[0], r[3]))

    available = {
        "ready": Sheet(tabs.get("ready", "آماده تولید محتوا"), READY_HEADER, ready),
        "archive": Sheet(tabs.get("archive", "آرشیو آینده"), ARCHIVE_HEADER, archive),
        "all": Sheet(tabs.get("all", "همه محصولات"), ALL_HEADER, everything),
        "review": Sheet(tabs.get("review", "نیاز به بازبینی دستی"), REVIEW_HEADER, review),
    }
    return [available[key] for key in wanted]


# ---------------------------------------------------------------------------
# نویسنده‌ها
# ---------------------------------------------------------------------------


class SheetsWriter:
    """نوشتن در Google Sheets با ``gspread`` — فقط یک batch در پایان."""

    def __init__(self, sheet_id: str, service_account_json: str) -> None:
        self.sheet_id = sheet_id
        self.service_account_json = service_account_json

    @property
    def available(self) -> bool:
        if not (self.sheet_id and self.service_account_json):
            return False
        try:  # pragma: no cover - وابسته به محیط
            import importlib

            importlib.import_module("gspread")
        except ImportError:
            return False
        return Path(self.service_account_json).exists()

    def write(self, sheets: Sequence[Sheet]) -> str:  # pragma: no cover - نیازمند شبکه
        import gspread  # type: ignore

        client = gspread.service_account(filename=self.service_account_json)
        spreadsheet = client.open_by_key(self.sheet_id)
        for sheet in sheets:
            values = [sheet.header] + [[_cell(v) for v in row] for row in sheet.rows]
            try:
                worksheet = spreadsheet.worksheet(sheet.title)
                worksheet.clear()
            except gspread.WorksheetNotFound:
                worksheet = spreadsheet.add_worksheet(
                    title=sheet.title,
                    rows=max(len(values) + 10, 100),
                    cols=max(len(sheet.header) + 2, 12),
                )
            worksheet.update(values, "A1", value_input_option="RAW")
        return f"https://docs.google.com/spreadsheets/d/{self.sheet_id}"


def write_xlsx(path: str | Path, sheets: Sequence[Sheet]) -> str:
    """خروجی محلی. اگر ``openpyxl`` نبود، به چند فایل CSV برمی‌گردیم."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        return _write_csv_bundle(path, sheets)

    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet in sheets:
        worksheet = workbook.create_sheet(title=sheet.title[:31])
        worksheet.sheet_view.rightToLeft = True
        worksheet.append(sheet.header)
        worksheet.freeze_panes = "A2"
        for row in sheet.rows:
            worksheet.append([_cell(v) for v in row])
        for index, column in enumerate(sheet.header, start=1):
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=index).column_letter
            ].width = max(18, min(60, len(str(column)) + 12))
    workbook.save(path)
    return str(path)


def sheet_to_csv(sheet: Sheet) -> bytes:
    """یک شیت → بایت‌های CSV با BOM، تا اکسل فارسی را درست باز کند."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(sheet.header)
    for row in sheet.rows:
        writer.writerow([_cell(value) for value in row])
    return buffer.getvalue().encode("utf-8-sig")


def sheets_to_xlsx(sheets: Sequence[Sheet]) -> bytes | None:
    """یک یا چند شیت → بایت‌های xlsx. بدون ``openpyxl`` خروجی ``None`` است."""
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        return None
    buffer = io.BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet in sheets:
        worksheet = workbook.create_sheet(title=sheet.title[:31])
        worksheet.sheet_view.rightToLeft = True
        worksheet.append(sheet.header)
        worksheet.freeze_panes = "A2"
        for row in sheet.rows:
            worksheet.append([_cell(value) for value in row])
        for index, column in enumerate(sheet.header, start=1):
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=index).column_letter
            ].width = max(18, min(60, len(str(column)) + 12))
    workbook.save(buffer)
    return buffer.getvalue()


def _write_csv_bundle(path: Path, sheets: Sequence[Sheet]) -> str:
    written: list[str] = []
    for sheet in sheets:
        target = path.with_name(f"{path.stem}__{_slug(sheet.title)}.csv")
        with target.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(sheet.header)
            for row in sheet.rows:
                writer.writerow([_cell(v) for v in row])
        written.append(str(target))
    return ", ".join(written)


def _cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return " | ".join(str(v) for v in value)
    if isinstance(value, bool):
        return "بله" if value else "خیر"
    return value


def _slug(text: str) -> str:
    return "".join(ch if ch.isalnum() else "-" for ch in text).strip("-") or "sheet"


# ---------------------------------------------------------------------------
# اجرا
# ---------------------------------------------------------------------------


def run(
    conn: sqlite3.Connection,
    run_id: str,
    config: Config,
    push_to_sheets: bool = True,
    verbose: bool = True,
) -> ExportStats:
    sheets = build_sheets(conn, run_id, config)
    counts = {key: len(sheet.rows) for key, sheet in zip(selected_sheets(config), sheets)}
    stats = ExportStats(
        ready_rows=counts.get("ready", 0),
        archive_rows=counts.get("archive", 0),
        all_rows=counts.get("all", 0),
        review_rows=counts.get("review", 0),
    )

    xlsx_path = config.get("output.xlsx_path", "content_pipeline/data/output-{run_id}.xlsx")
    if xlsx_path:
        stats.local_path = write_xlsx(str(xlsx_path).replace("{run_id}", run_id), sheets)

    if push_to_sheets:
        writer = SheetsWriter(
            config.get("output.sheet_id", ""), config.get("output.service_account_json", "")
        )
        if writer.available:
            stats.sheet_url = writer.write(sheets)
        elif verbose:
            print("  گوگل‌شیت تنظیم نشده است؛ فقط خروجی محلی ساخته شد.")

    return stats
